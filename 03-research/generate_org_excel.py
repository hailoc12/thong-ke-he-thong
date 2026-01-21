#!/usr/bin/env python3
"""
Generate Excel file with organization accounts
Created: 2026-01-21
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re

# 34 organizations from Word doc
ORGANIZATIONS = [
    "Vụ Bưu chính",
    "Vụ Đánh giá và Thẩm định công nghệ",
    "Vụ Khoa học kỹ thuật và công nghệ",
    "Vụ Khoa học Xã hội, Nhân văn và Tự nhiên",
    "Vụ Kinh tế và Xã hội số",
    "Văn phòng Bộ",
    "Cục An toàn bức xạ và hạt nhân",
    "Cục Bưu điện điện Trung ương",
    "Cục Công nghiệp Công nghệ thông tin",
    "Cục Chuyển đổi số quốc gia",
    "Cục Đổi mới sáng tạo",
    "Cục Khởi nghiệp và Doanh nghiệp công nghệ",
    "Cục Sở hữu trí tuệ",
    "Cục Tần số vô tuyến điện",
    "Cục Thông tin, Thống kê",
    "Cục Viễn thông",
    "Uỷ ban Tiêu chuẩn Đo lường Chất lượng quốc gia",
    "Trung tâm Công nghệ thông tin",
    "Học viện Công nghệ Bưu chính Viễn thông",
    "Học viện Chiến lược Khoa học và Công nghệ",
    "Báo VNExpress",
    "Quỹ Phát triển khoa học và công nghệ quốc gia",
    "Quỹ Đổi mới công nghệ quốc gia",
    "Quỹ Dịch vụ viễn thông công ích Việt Nam",
    "Viện Công nghệ số và Chuyển đổi số quốc gia",
    "Viện Năng lượng nguyên tử Việt Nam",
    "Viện Khoa học và Công nghệ Việt Nam - Hàn Quốc",
    "Viện Sở hữu trí tuệ quốc gia",
    "Viện Ứng dụng công nghệ",
    "Trung tâm Chứng thực điện tử quốc gia",
    "Trung tâm Internet Việt Nam",
    "Trung tâm Truyền thông khoa học và công nghệ",
    "Nhà Xuất bản Khoa học - Công nghệ - Truyền thông",
    "Trường Cao đẳng Thông tin và Truyền thông",
]

PASSWORD = "ThongkeCDS@2026#"

def generate_username(org_name):
    """
    Generate username from organization name following pattern:
    - Vụ Bưu chính → vu-buuchinh
    - Cục An toàn bức xạ → cuc-atbx (abbreviation)
    - Học viện Công nghệ Bưu chính Viễn thông → ptit (well-known abbreviation)
    """

    # Special cases with well-known abbreviations
    special_cases = {
        "Học viện Công nghệ Bưu chính Viễn thông": "ptit",
        "Báo VNExpress": "vnexpress",
        "Cục An toàn bức xạ và hạt nhân": "cuc-atbxhn",
        "Uỷ ban Tiêu chuẩn Đo lường Chất lượng quốc gia": "stameq",
        "Viện Năng lượng nguyên tử Việt Nam": "vinatom",
        "Trung tâm Internet Việt Nam": "vnnic",
        "Học viện Chiến lược Khoa học và Công nghệ": "nais",
    }

    if org_name in special_cases:
        return special_cases[org_name]

    # Remove Vietnamese tones
    tone_map = {
        'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        'đ': 'd', 'Đ': 'd',
    }

    # Normalize
    normalized = org_name.lower()
    for vn_char, latin_char in tone_map.items():
        normalized = normalized.replace(vn_char, latin_char)

    # Extract prefix (Vụ, Cục, Viện, etc.) and main part
    parts = normalized.split()

    if not parts:
        return "unknown"

    prefix = parts[0]  # Vụ, Cục, Viện, etc.
    main_parts = parts[1:]

    # Generate abbreviation from main parts
    # Take first letter of each word, skip common words
    skip_words = {'và', 'cua', 'cho', 'tren', 'duoi', 'trong', 'ngoai', 'voi'}

    # For longer names, create abbreviation
    if len(main_parts) >= 3:
        abbrev = ''.join([word[0] for word in main_parts if word not in skip_words])
        return f"{prefix}-{abbrev}"
    else:
        # For shorter names, use full words
        main = ''.join(main_parts)
        return f"{prefix}-{main}"

def create_excel():
    """Create Excel file with organization accounts"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách tài khoản"

    # Headers
    headers = ["Tên đơn vị", "Username", "Password"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add organization data
    for org_name in ORGANIZATIONS:
        username = generate_username(org_name)
        ws.append([org_name, username, PASSWORD])

    # Adjust column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25

    # Center align username and password columns
    for row in range(2, len(ORGANIZATIONS) + 2):
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')

    # Save file
    output_file = "/Users/shimazu/Dropbox/9. active/consultant/support_b4t/thong_ke_he_thong/03-research/danh-sach-tai-khoan-don-vi-NEW.xlsx"
    wb.save(output_file)

    print(f"✓ Created Excel file: {output_file}")
    print(f"✓ Total organizations: {len(ORGANIZATIONS)}")

    # Print sample usernames for verification
    print("\nSample usernames generated:")
    for i, org_name in enumerate(ORGANIZATIONS[:10], 1):
        username = generate_username(org_name)
        print(f"  {i}. {org_name:50s} → {username}")
    print("  ...")

if __name__ == "__main__":
    create_excel()
