#!/usr/bin/env python3
"""
Generate sample Excel file for CDS Dashboard Export
Format: Times New Roman 14, Colors, Multi-level headers
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# Colors
HEADER_DARK = "1F4E78"
HEADER_LIGHT = "D9E1F2"
HIGHLIGHT_YELLOW = "FFC000"
HIGHLIGHT_RED = "FFC7CE"
HIGHLIGHT_ORANGE = "FFEB9C"
HIGHLIGHT_LIGHT_YELLOW = "FFF2CC"
GREEN_BG = "C6EFCE"

# Font
FONT_HEADER = Font(name='Times New Roman', size=14, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name='Times New Roman', size=13, bold=True)
FONT_NORMAL = Font(name='Times New Roman', size=14)
FONT_TITLE = Font(name='Times New Roman', size=18, bold=True, color="FFFFFF")

# Alignment
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# Border
THIN_BORDER = Border(
    left=Side(style='thin', color="1F4E78"),
    right=Side(style='thin', color="1F4E78"),
    top=Side(style='thin', color="1F4E78"),
    bottom=Side(style='thin', color="1F4E78")
)

HBORDER = Border(
    bottom=Side(style='thin', color="1F4E78")
)

# Fills
FILL_DARK = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=HEADER_LIGHT, end_color=HEADER_LIGHT, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=HIGHLIGHT_YELLOW, end_color=HIGHLIGHT_YELLOW, fill_type="solid")
FILL_RED = PatternFill(start_color=HIGHLIGHT_RED, end_color=HIGHLIGHT_RED, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=HIGHLIGHT_ORANGE, end_color=HIGHLIGHT_ORANGE, fill_type="solid")
FILL_LIGHT_YELLOW = PatternFill(start_color=HIGHLIGHT_LIGHT_YELLOW, end_color=HIGHLIGHT_LIGHT_YELLOW, fill_type="solid")
FILL_GREEN = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")


def create_sheet1_summary(wb):
    """Sheet 1: Tong quan"""
    ws = wb.create_sheet("1. Tổng quan")

    # Title
    ws.merge_cells('A1:C1')
    cell = ws['A1']
    cell.value = "BÁO CÁO TỔNG QUAN HỆ THỐNG CDS"
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER
    cell.fill = FILL_DARK

    # Date
    ws.merge_cells('A2:C2')
    cell = ws['A2']
    cell.value = f"Ngày {datetime.now().strftime('%d/%m/%Y')}"
    cell.font = Font(name='Times New Roman', size=14, italic=True)
    cell.alignment = ALIGN_RIGHT

    # Header
    headers = ['CHỈ TIÊU', 'GIÁ TRỊ', 'GHI CHÚ']
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=idx, value=h)
        cell.font = FONT_HEADER
    ws['A4'].fill = FILL_DARK
    ws['B4'].fill = FILL_DARK
    ws['C4'].fill = FILL_DARK
    ws['A4'].alignment = ALIGN_CENTER
    ws['B4'].alignment = ALIGN_CENTER
    ws['C4'].alignment = ALIGN_CENTER

    # Data
    data = [
        ['THEO TRẠNG THÁI', '', ''],
        ['1. Tổng số hệ thống', '125', ''],
        ['2. Số hệ thống đang hoạt động', '98', '78.4%'],
        ['3. Số hệ thống thí điểm', '15', '12.0%'],
        ['4. Số hệ thống dừng', '8', '6.4%'],
        ['5. Số hệ thống sắp thay thế', '4', '3.2%'],
        ['', '', ''],
        ['THEO MỨC ĐỘ QUAN TRỌNG', '', ''],
        ['6. Cực kỳ quan trọng', '35', '28.0%'],
        ['7. Quan trọng', '50', '40.0%'],
        ['8. Trung bình', '30', '24.0%'],
        ['9. Thấp', '10', '8.0%'],
        ['', '', ''],
        ['THEO TRÌNH ĐỘ NHẬP LIỆU', '', ''],
        ['10. Tỷ lệ nhập liệu TB', '65.8%', ''],
        ['11. Số hệ thống hoàn thành 100%', '42', '33.6%'],
        ['12. Số hệ thống dưới 50%', '25', '20.0%'],
    ]

    row_idx = 5
    for row_data in data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_NORMAL
            cell.border = HBORDER
            if col_idx == 1 and row_data[0].startswith('THEO'):
                cell.font = Font(name='Times New Roman', size=14, bold=True)
                cell.fill = FILL_LIGHT
            if col_idx == 2:
                cell.alignment = ALIGN_RIGHT
        row_idx += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_sheet2_orgs(wb):
    """Sheet 2: Thong ke theo don vi"""
    ws = wb.create_sheet("2. Theo đơn vị")

    # Single header row - simplified format
    from datetime import datetime, timedelta
    # Vietnam time (UTC+7)
    vn_now = datetime.now() + timedelta(hours=7)
    time_str = vn_now.strftime('%d/%m %H:%M')

    # Row 1 - Main header with merged cells
    # Merge D1:E1 for the percentage column header
    ws.merge_cells('D1:E1')
    cell = ws.cell(row=1, column=1, value='STT')
    cell.font = FONT_HEADER
    cell.fill = FILL_DARK
    cell.alignment = ALIGN_CENTER

    cell = ws.cell(row=1, column=2, value='ĐƠN VỊ')
    cell.font = FONT_HEADER
    cell.fill = FILL_DARK
    cell.alignment = ALIGN_CENTER

    cell = ws.cell(row=1, column=3, value='TRẠNG THÁI')
    cell.font = FONT_HEADER
    cell.fill = FILL_DARK
    cell.alignment = ALIGN_CENTER

    cell = ws.cell(row=1, column=4, value=f'TỶ LỆ NHẬP DỮ LIỆU TÍNH ĐẾN {time_str}')
    cell.font = FONT_HEADER
    cell.fill = FILL_DARK
    cell.alignment = ALIGN_CENTER

    # Row 2 - Sub headers
    ws.cell(row=2, column=4, value='Số hệ thống').font = FONT_SUBHEADER
    ws.cell(row=2, column=4).fill = FILL_LIGHT
    ws.cell(row=2, column=4).alignment = ALIGN_CENTER

    ws.cell(row=2, column=5, value='% hoàn thành trung bình').font = FONT_SUBHEADER
    ws.cell(row=2, column=5).fill = FILL_LIGHT
    ws.cell(row=2, column=5).alignment = ALIGN_CENTER

    # Dummy data - simplified format (STT, Ten, Trang thai, So HT, % HT)
    orgs = [
        ('Trường Cao đẳng Thông tin và Truyền thông', 'Hoạt động', 2, 26.1),
        ('Nhà Xuất bản Khoa học - Công nghệ - Truyền thông', 'Hoạt động', 0, 0),
        ('Trung tâm Internet Việt Nam', 'Hoạt động', 1, 18.8),
        ('Trung tâm Chứng thực điện tử quốc gia', 'Hoạt động', 7, 50.43),
        ('Viện Ứng dụng công nghệ', 'Hoạt động', 0, 0),
        ('Viện Sở hữu trí tuệ quốc gia', 'Hoạt động', 1, 65.30),
        ('Viện Khoa học và Công nghệ - Hàn Quốc', 'Hoạt động', 1, 65.30),
        ('Viện Năng lượng nguyên tử Việt Nam', 'Hoạt động', 0, 0),
        ('Viện Công nghệ số và Chuyển đổi số quốc gia', 'Hoạt động', 1, 57.10),
        ('Quỹ Đổi mới công nghệ quốc gia', 'Hoạt động', 0, 0),
        ('Quỹ Phát triển khoa học và công nghệ quốc gia', 'Hoạt động', 0, 0),
        ('Báo VNEExpress', 'Hoạt động', 1, 80.40),
        ('Báo Văn hóa - Văn nghệ Khoa học và Công nghệ', 'Hoạt động', 1, 63.30),
        ('Học viện Công nghệ Bưu chính Viễn thông', 'Hoạt động', 3, 56.47),
        ('Trung tâm Công nghệ thông tin', 'Hoạt động', 26, 68.78),
        ('Uỷ ban Tiêu chuẩn Đo lường Chất lượng quốc gia', 'Hoạt động', 6, 50.97),
        ('Cục Viễn thông', 'Hoạt động', 1, 55.10),
        ('Cục Thông tin, Thống kê', 'Hoạt động', 5, 68.14),
        ('Cục Tần số vô tuyến điện', 'Hoạt động', 5, 62.86),
        ('Cục Sở hữu trí tuệ', 'Hoạt động', 0, 0),
        ('Cục Khởi nghiệp và Doanh nghiệp công nghệ', 'Hoạt động', 0, 0),
        ('Cục Điện tử công nghệ', 'Hoạt động', 0, 0),
        ('Cục Chuyển đổi số quốc gia', 'Hoạt động', 2, 55.10),
        ('Cục Công nghiệp Công nghệ thông tin', 'Hoạt động', 0, 0),
        ('Cục Bưu điện điện Trung ương', 'Hoạt động', 2, 66.35),
        ('Cục An toàn bức xạ và hạt nhân', 'Hoạt động', 2, 35.75),
        ('Văn phòng Bộ', 'Hoạt động', 1, 26.50),
        ('Vụ Kinh tế và Xã hội số', 'Hoạt động', 1, 22.40),
        ('Vụ Khoa học Xã hội, Nhân văn và Tự nhiên', 'Hoạt động', 0, 0),
        ('Vụ Khoa học kỹ thuật và công nghệ', 'Hoạt động', 0, 0),
        ('Vụ Đánh giá và Thẩm định công nghệ', 'Hoạt động', 1, 53.10),
        ('Vụ Bưu chính', 'Hoạt động', 2, 74.50),
    ]

    row = 3
    for idx, (name, status, count, pct) in enumerate(orgs, 1):
        ws.cell(row=row, column=1, value=idx).font = FONT_NORMAL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER

        ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT

        ws.cell(row=row, column=3, value=status).font = FONT_NORMAL
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = ALIGN_CENTER

        cell = ws.cell(row=row, column=4, value=count if count > 0 else 'Chưa có dữ liệu')
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER

        cell = ws.cell(row=row, column=5, value=f"{pct:.2f}".replace('.', ',') if pct > 0 else 'Chưa có dữ liệu')
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER

        row += 1

    # Summary section - like the image
    row += 1

    # Calculate summary stats
    total_orgs = len(orgs)
    updated = sum(1 for _, _, _, pct in orgs if pct > 0)
    not_updated = total_orgs - updated

    above_80 = sum(1 for _, _, _, pct in orgs if pct >= 80)
    range_60_80 = sum(1 for _, _, _, pct in orgs if 60 <= pct < 80)
    below_30 = sum(1 for _, _, _, pct in orgs if 0 < pct < 30)
    below_60 = sum(1 for _, _, _, pct in orgs if 0 < pct < 60)

    # Merge for total count
    ws.cell(row=row, column=1, value=f'Tổng hợp: {total_orgs} Đơn vị')
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1)
    cell.font = FONT_HEADER
    cell.fill = FILL_YELLOW
    cell.alignment = ALIGN_LEFT

    row += 1
    ws.cell(row=row, column=1, value='Đã cập nhật')
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name='Times New Roman', size=14, bold=True)
    cell.alignment = ALIGN_LEFT

    row += 1
    ws.cell(row=row, column=1, value='Trong đó')
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name='Times New Roman', size=14, bold=True)
    cell.alignment = ALIGN_LEFT

    row += 1
    # Summary items with format: "Tỷ lệ CNDL > 80%" | [count]
    ws.cell(row=row, column=2, value=f'Tỷ lệ CNDL > 80%')
    ws.cell(row=row, column=5, value=str(above_80))
    ws.cell(row=row, column=2).font = FONT_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_LEFT
    ws.cell(row=row, column=5).font = FONT_NORMAL
    ws.cell(row=row, column=5).alignment = ALIGN_CENTER

    row += 1
    ws.cell(row=row, column=2, value=f'Tỷ lệ CNDL từ 60% đến 80%')
    ws.cell(row=row, column=5, value=str(range_60_80))
    ws.cell(row=row, column=2).font = FONT_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_LEFT
    ws.cell(row=row, column=5).font = FONT_NORMAL
    ws.cell(row=row, column=5).alignment = ALIGN_CENTER

    row += 1
    ws.cell(row=row, column=2, value=f'Tỷ lệ CNDL < 60%')
    ws.cell(row=row, column=5, value=str(below_60))
    ws.cell(row=row, column=2).font = FONT_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_LEFT
    ws.cell(row=row, column=5).font = FONT_NORMAL
    ws.cell(row=row, column=5).alignment = ALIGN_CENTER

    row += 1
    ws.cell(row=row, column=2, value=f'Tỷ lệ CNDL <30%')
    ws.cell(row=row, column=5, value=str(below_30))
    ws.cell(row=row, column=2).font = FONT_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_LEFT
    ws.cell(row=row, column=5).font = FONT_NORMAL
    ws.cell(row=row, column=5).alignment = ALIGN_CENTER

    row += 1
    ws.cell(row=row, column=1, value='Chưa cập nhật')
    ws.cell(row=row, column=5, value=str(not_updated))
    ws.cell(row=row, column=1).font = Font(name='Times New Roman', size=14, bold=True)
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.cell(row=row, column=5).font = Font(name='Times New Roman', size=14, bold=True)
    ws.cell(row=row, column=5).alignment = ALIGN_CENTER

    # List of units not updated
    row += 2
    ws.cell(row=row, column=1, value='Danh sách đơn vị chưa cập nhật dữ liệu')
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws.cell(row=row, column=1)
    cell.font = FONT_HEADER
    cell.fill = FILL_DARK
    cell.alignment = ALIGN_LEFT

    not_updated_orgs = [name for name, _, _, pct in orgs if pct == 0]
    for org_name in not_updated_orgs:
        row += 1
        ws.cell(row=row, column=1, value=org_name)
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

    # Freeze rows
    ws.freeze_panes = 'A3'


def create_sheet3_systems(wb):
    """Sheet 3: Danh sach he thong"""
    ws = wb.create_sheet("3. Danh sách HT")

    # Header
    headers = ['STT', 'TÊN HỆ THỐNG', 'ĐƠN VỊ', 'TRẠNG THÁI', 'QUAN TRỌNG', '% HOÀN THÀNH', 'NGÀY CẬP NHẬT']
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_DARK
        cell.alignment = ALIGN_CENTER

    # Dummy systems - sorted by org, then completion desc
    systems = [
        ['Hệ thống quản lý tài sản CNTT', 'Cục Công nghệ thông tin', 'Hoạt động', 'Cực kỳ quan trọng', 85.5, '25/01/2026'],
        ['Hệ thống giám sát an toàn thông tin', 'Cục Công nghệ thông tin', 'Hoạt động', 'Quan trọng', 72.3, '24/01/2026'],
        ['Hệ thống quản lý người dùng', 'Cục Công nghệ thông tin', 'Thí điểm', 'Trung bình', 45.0, '20/01/2026'],
        ['Hệ thống báo cáo thống kê', 'Trung tâm Internet Việt Nam', 'Hoạt động', 'Thấp', 92.0, '25/01/2026'],
        ['Hệ thống DNS Việt Nam', 'Trung tâm Internet Việt Nam', 'Hoạt động', 'Quan trọng', 24.5, '15/01/2026'],
        ['Hệ thống chứng thực điện tử', 'Trung tâm Chứng thực điện tử QG', 'Hoạt động', 'Cực kỳ quan trọng', 100.0, '26/01/2026'],
        ['Hệ thống bảo mật mạng', 'Trung tâm Chứng thực điện tử QG', 'Hoạt động', 'Quan trọng', 88.5, '25/01/2026'],
        ['Hệ thống quản lý mã số', 'Viện Sở hữu trí tuệ QG', 'Hoạt động', 'Trung bình', 65.3, '23/01/2026'],
        ['Hệ thống portal công khai', 'Viện Sở hữu trí tuệ QG', 'Hoạt động', 'Thấp', 100.0, '26/01/2026'],
        ['Hệ thống đào tạo trực tuyến', 'Trường CĐ TT&TT', 'Hoạt động', 'Trung bình', 26.1, '18/01/2026'],
        ['Hệ thống quản lý sinh viên', 'Trường CĐ TT&TT', 'Hoạt động', 'Quan trọng', 63.3, '24/01/2026'],
    ]

    row = 2
    for idx, sys_data in enumerate(systems, 1):
        ws.cell(row=row, column=1, value=idx).font = FONT_NORMAL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER

        # Check completion percentage for color coding (column F, index 6)
        # sys_data[4] is the completion percentage value
        pct = float(sys_data[4]) if sys_data[4] is not None else 0.0

        for col_idx, value in enumerate(sys_data, 2):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER

            # Alignment
            if col_idx in [2, 3]:  # Ten HT, Don vi
                cell.alignment = ALIGN_LEFT
            else:
                cell.alignment = ALIGN_CENTER

            # Color code completion percentage
            if col_idx == 6:  # Completion column
                if pct >= 80:
                    cell.fill = FILL_GREEN
                elif pct >= 60:
                    cell.fill = FILL_LIGHT_YELLOW
                else:
                    cell.fill = FILL_RED
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    # Freeze header
    ws.freeze_panes = 'A2'


def create_sheet4_followup(wb):
    """Sheet 4: Luu y don doc"""
    ws = wb.create_sheet("4. Lưu ý đôn đốc")

    # Header
    headers = ['STT', 'ĐƠN VỊ', '% HOÀN THÀNH', 'LƯU Ý / GHI CHÚ']
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_DARK
        cell.alignment = ALIGN_CENTER

    # Orgs needing follow-up (0-50% completion)
    followup = [
        ('Nhà Xuất bản KHCN - TT&TT', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Viện Ứng dụng công nghệ', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Viện Năng lượng nguyên tử VN', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Quỹ Đổi mới công nghệ QG', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Cục Công nghiệp CNTT', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Cục Sở hữu trí tuệ', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Vụ Khoa học Kỹ thuật', 0, 'Chưa cập nhật dữ liệu', 'red'),
        ('Trường CĐ TT&TT', 26.1, 'Cần đốc thúc nhập liệu', 'orange'),
        ('Trung tâm Internet VN', 18.8, 'Cần đốc thúc nhập liệu', 'orange'),
        ('Cục An toàn bức xạ', 23.9, 'Cần đốc thúc nhập liệu', 'orange'),
        ('Văn phòng Bộ', 17.4, 'Cần đốc thúc nhập liệu', 'orange'),
        ('Cục Bưu điện Trung ương', 31.1, 'Cần quan tâm跟进', 'yellow'),
        ('Vụ Kinh tế và Xã hội số', 35.2, 'Cần quan tâm跟进', 'yellow'),
        ('Vụ Khoa học Xã hội', 42.5, 'Tiếp tục monitor', 'yellow'),
    ]

    row = 2
    for idx, (org_name, pct, note, color) in enumerate(followup, 1):
        ws.cell(row=row, column=1, value=idx).font = FONT_NORMAL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER

        ws.cell(row=row, column=2, value=org_name).font = FONT_NORMAL
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT

        cell = ws.cell(row=row, column=3, value=f"{pct}%")
        cell.font = FONT_NORMAL
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER

        if color == 'red':
            cell.fill = FILL_RED
        elif color == 'orange':
            cell.fill = FILL_ORANGE
        else:
            cell.fill = FILL_LIGHT_YELLOW

        ws.cell(row=row, column=4, value=note).font = FONT_NORMAL
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = ALIGN_LEFT

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40

    # Freeze header
    ws.freeze_panes = 'A2'


def main():
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    create_sheet1_summary(wb)
    create_sheet2_orgs(wb)
    create_sheet3_systems(wb)
    create_sheet4_followup(wb)

    filename = f"Bao-cao-CDS-mau-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    wb.save(filename)
    print(f"Created: {filename}")
    print("\nSheets:")
    print("  1. Tổng quan - Summary statistics")
    print("  2. Theo đơn vị - Organization breakdown")
    print("  3. Danh sách HT - Systems list")
    print("  4. Lưu ý đôn đốc - Follow-up list")


if __name__ == '__main__':
    main()
