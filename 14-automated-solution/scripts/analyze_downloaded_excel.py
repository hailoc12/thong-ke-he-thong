#!/usr/bin/env python3
"""Analyze downloaded Excel file"""

from openpyxl import load_workbook
import sys

excel_path = ".playwright-mcp/Bao-cao-CDS-26-01-2026.xlsx"

try:
    print("📊 Analyzing Downloaded Excel File...")
    print("=" * 60)

    wb = load_workbook(excel_path)

    print(f"File: {excel_path}")
    print(f"Total sheets: {len(wb.sheetnames)}\n")

    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_count = ws.max_row
        col_count = ws.max_column
        results[sheet_name] = row_count

        # Get header row
        headers = []
        if row_count >= 1:
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))

        print(f"📄 Sheet: '{sheet_name}'")
        print(f"   Rows: {row_count} (including header)")
        print(f"   Cols: {col_count}")
        if headers:
            print(f"   Headers: {', '.join(headers[:5])}")
        print()

    print("=" * 60)
    print("VERIFICATION RESULTS:")
    print("=" * 60)

    # Check Sheet 2
    sheet2_name = '2. Theo đơn vị'
    sheet2_rows = results.get(sheet2_name, 0)
    if sheet2_rows >= 33:
        print(f"✅ Sheet 2 ({sheet2_name}): {sheet2_rows} rows")
        print(f"   Expected: ≥33 (32 organizations + header)")
        print(f"   Result: PASS - {sheet2_rows - 1} organizations included")
    else:
        print(f"❌ Sheet 2 ({sheet2_name}): {sheet2_rows} rows")
        print(f"   Expected: ≥33, Got: {sheet2_rows}")
        print(f"   Result: FAIL")
    print()

    # Check Sheet 3
    sheet3_name = '3. Danh sách HT'
    sheet3_rows = results.get(sheet3_name, 0)
    if sheet3_rows >= 78:
        print(f"✅ Sheet 3 ({sheet3_name}): {sheet3_rows} rows")
        print(f"   Expected: ≥78 (77 systems + header)")
        print(f"   Result: PASS - All {sheet3_rows - 1} systems included")
        print(f"   🎉 FIX IS WORKING!")
    elif sheet3_rows == 21:
        print(f"❌ Sheet 3 ({sheet3_name}): {sheet3_rows} rows")
        print(f"   Expected: ≥78, Got: {sheet3_rows}")
        print(f"   Result: FAIL - Bug still present (only 20 systems)")
    else:
        print(f"⚠️  Sheet 3 ({sheet3_name}): {sheet3_rows} rows")
        print(f"   Expected: ≥78, Got: {sheet3_rows}")
        print(f"   Result: Partial data ({sheet3_rows - 1} systems)")

    print("=" * 60)

    if sheet2_rows >= 33 and sheet3_rows >= 78:
        print("🎉 ALL TESTS PASSED - Excel export working correctly!")
        sys.exit(0)
    else:
        print("❌ TESTS FAILED - Excel export still has issues")
        sys.exit(1)

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
