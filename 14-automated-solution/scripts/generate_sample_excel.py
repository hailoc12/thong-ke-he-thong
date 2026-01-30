#!/usr/bin/env python3
"""
Generate sample Excel file for System Export feature
FULL VERSION - All fields from SystemCreate/SystemEdit forms
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import random

# Styling constants
HEADER_FILL = PatternFill(start_color="1890ff", end_color="1890ff", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='d9d9d9'),
    right=Side(style='thin', color='d9d9d9'),
    top=Side(style='thin', color='d9d9d9'),
    bottom=Side(style='thin', color='d9d9d9')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Status colors
STATUS_COLORS = {
    'Đang vận hành': PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid"),
    'Pilot': PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid"),
    'Thử nghiệm': PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid"),
    'Dừng': PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid"),
    'Sắp thay thế': PatternFill(start_color="ffeeba", end_color="ffeeba", fill_type="solid"),
}

CRITICALITY_FONTS = {
    'Quan trọng': Font(color="dc3545", bold=True),
    'Trung bình': Font(color="fd7e14"),
    'Thấp': Font(color="28a745"),
}

def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

def apply_data_style(ws, start_row=2):
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

def auto_fit_columns(ws, min_width=8, max_width=50):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def bool_vn(val):
    return 'Có' if val else 'Không'

def list_to_str(lst):
    if isinstance(lst, list):
        return ', '.join(str(x) for x in lst)
    return str(lst) if lst else ''

def date_str(d):
    if d:
        return d.strftime('%d/%m/%Y')
    return ''

def num_fmt(n):
    if n is None:
        return ''
    return f"{n:,}"

# Sample data generators
ORGS = ["Sở TTTT", "Sở Tài chính", "Sở KH&ĐT", "Sở Nội vụ", "VP UBND"]
STATUSES = ["Đang vận hành", "Pilot", "Thử nghiệm", "Dừng", "Sắp thay thế"]
CRITICALITIES = ["Quan trọng", "Trung bình", "Thấp"]
SCOPES = ["Nội bộ đơn vị", "Toàn bộ", "Bên ngoài"]
GROUPS = ["Nền tảng quốc gia", "Nền tảng dùng chung", "CSDL chuyên ngành", "Ứng dụng nghiệp vụ", "Cổng thông tin"]
REQ_TYPES = ["Xây mới", "Nâng cấp", "Tích hợp", "Thay thế", "Mở rộng"]
ARCH_TYPES = ["Monolithic", "Modular", "Microservices", "SOA", "Serverless"]
DATABASES = ["PostgreSQL", "MySQL", "Oracle", "SQL Server", "MongoDB"]
HOSTING = ["Cloud", "On-premise", "Hybrid"]
DEV_TYPES = ["Nội bộ", "Thuê ngoài", "Kết hợp"]
MOBILE = ["Native App", "Hybrid App", "PWA", "Không có"]
DB_MODEL = ["Tập trung", "Phân tán", "Riêng từng app"]
CONTAINER = ["Docker", "Kubernetes", "OpenShift", "Không sử dụng"]
API_STYLE = ["REST API", "GraphQL", "gRPC", "SOAP"]
MSG_QUEUE = ["Kafka", "RabbitMQ", "ActiveMQ", "Không sử dụng"]
CACHE = ["Redis", "Memcached", "Không sử dụng"]
SEARCH = ["Elasticsearch", "Solr", "Không sử dụng"]
BI_TOOL = ["Power BI", "Tableau", "Metabase", "Superset", "Tự phát triển", "Không có"]
SOURCE_REPO = ["GitLab", "GitHub", "Bitbucket", "Azure DevOps"]
CICD = ["Jenkins", "GitLab CI", "GitHub Actions", "Azure Pipelines"]
FILE_STORAGE = ["File Server", "Object Storage", "NAS", "Database BLOB"]
WARRANTY = ["Còn bảo hành", "Hết bảo hành", "Không có"]
VENDOR_DEP = ["Cao", "Trung bình", "Thấp", "Không"]
DEPLOY_LOC = ["Data Center", "Cloud", "Hybrid"]
COMPUTE_TYPE = ["Virtual Machine", "Container", "Serverless", "Bare Metal"]
DEPLOY_FREQ = ["Daily", "Weekly", "Monthly", "Quarterly", "On Demand"]
API_STD = ["OpenAPI 3.0", "Swagger", "SOAP WSDL", "GraphQL Schema"]
RECOMMEND = ["Giữ nguyên", "Nâng cấp", "Thay thế", "Hợp nhất"]

def generate_sample_data(n=5):
    systems = []
    for i in range(1, n + 1):
        sys = {
            # Tab 1: Cơ bản
            'stt': i,
            'code': f'SYS-STTTT-2024-{str(i).zfill(4)}',
            'name': f'Hệ thống {["Quản lý văn bản", "Nhân sự", "Tài chính", "Dự án", "GIS"][i-1]}',
            'name_en': f'System {["Document", "HR", "Finance", "Project", "GIS"][i-1]}',
            'org': ORGS[i % len(ORGS)],
            'purpose': 'Hỗ trợ quản lý, điều hành và nâng cao hiệu quả công việc',
            'status': STATUSES[i % len(STATUSES)],
            'criticality': CRITICALITIES[i % len(CRITICALITIES)],
            'scope': SCOPES[i % len(SCOPES)],
            'req_type': REQ_TYPES[i % len(REQ_TYPES)],
            'target_date': date(2025, (i * 3) % 12 + 1, 15),
            'system_group': GROUPS[i % len(GROUPS)],
            'notes_tab1': 'Ghi chú tab cơ bản',

            # Tab 2: Nghiệp vụ
            'biz_objectives': ['Số hóa quy trình', 'Cải thiện dịch vụ công', 'Tăng cường minh bạch'],
            'biz_processes': ['Quản lý hồ sơ', 'Phê duyệt', 'Tra cứu', 'Báo cáo'],
            'has_design_docs': random.choice([True, False]),
            'user_types': ['Lãnh đạo nội bộ', 'Cán bộ nội bộ', 'Doanh nghiệp', 'Người dân'],
            'annual_users': random.randint(1000, 50000),
            'total_accounts': random.randint(100, 5000),
            'mau': random.randint(100, 2000),
            'dau': random.randint(20, 500),
            'num_orgs': random.randint(1, 50),
            'notes_tab2': 'Ghi chú tab nghiệp vụ',

            # Tab 3: Kiến trúc
            'lang': random.choice(['Python, JavaScript', 'Java, TypeScript', 'C#, JavaScript']),
            'framework': random.choice(['Django, React', 'Spring Boot, Vue.js', '.NET, Angular']),
            'database': DATABASES[i % len(DATABASES)],
            'hosting': HOSTING[i % len(HOSTING)],
            'backend_tech': random.choice(['Python', 'Java', 'C#/.NET', 'Node.js']),
            'frontend_tech': random.choice(['React', 'Vue.js', 'Angular']),
            'arch_type': ARCH_TYPES[i % len(ARCH_TYPES)],
            'containerization': CONTAINER[i % len(CONTAINER)],
            'is_multi_tenant': random.choice([True, False]),
            'has_layered': random.choice([True, False]),
            'layered_details': 'Presentation, Business Logic, Data Access',
            'api_style': API_STYLE[i % len(API_STYLE)],
            'msg_queue': MSG_QUEUE[i % len(MSG_QUEUE)],
            'cache': CACHE[i % len(CACHE)],
            'search': SEARCH[i % len(SEARCH)],
            'bi_tool': BI_TOOL[i % len(BI_TOOL)],
            'source_repo': SOURCE_REPO[i % len(SOURCE_REPO)],
            'has_cicd': random.choice([True, False]),
            'cicd_tool': CICD[i % len(CICD)],
            'has_auto_test': random.choice([True, False]),
            'test_tools': 'Jest, Pytest, Selenium',
            'notes_tab3': 'Ghi chú tab kiến trúc',

            # Tab 4: Dữ liệu
            'data_sources': ['User Input', 'External APIs', 'Database Sync', 'File Import'],
            'data_types': ['Business Data', 'Documents', 'Statistics'],
            'data_class': random.choice(['Công khai', 'Nội bộ', 'Hạn chế', 'Bí mật']),
            'data_volume': f'{random.randint(1, 100)} GB',
            'storage_gb': random.randint(10, 500),
            'file_storage_gb': random.randint(5, 200),
            'growth_rate': random.randint(5, 30),
            'file_storage_type': FILE_STORAGE[i % len(FILE_STORAGE)],
            'record_count': random.randint(10000, 1000000),
            'secondary_dbs': random.choice(['Redis', 'MongoDB', 'Không']),
            'has_catalog': random.choice([True, False]),
            'catalog_notes': 'Đang xây dựng data catalog',
            'has_mdm': random.choice([True, False]),
            'mdm_notes': 'Master data cho nhân sự',
            'notes_tab4': 'Ghi chú tab dữ liệu',

            # Tab 5: Tích hợp
            'api_provided': random.randint(0, 20),
            'api_consumed': random.randint(0, 15),
            'api_standard': API_STD[i % len(API_STD)],
            'has_api_gateway': random.choice([True, False]),
            'gateway_name': random.choice(['Kong', 'AWS API Gateway', 'Nginx']),
            'has_versioning': random.choice([True, False]),
            'has_rate_limit': random.choice([True, False]),
            'api_docs': 'https://api.example.gov.vn/docs',
            'version_std': 'Semantic Versioning (v1.2.3)',
            'has_int_monitor': random.choice([True, False]),
            'internal_systems': ['Hệ thống văn bản', 'Portal nội bộ', 'SSO'],
            'external_systems': ['LGSP', 'VNeID', 'Cổng DVCQG'],
            'api_list': ['GET /users', 'POST /documents', 'PUT /approvals'],
            'exchange_method': random.choice(['API REST', 'File Transfer', 'Message Queue']),
            'notes_tab5': 'Ghi chú tab tích hợp',

            # Tab 6: Bảo mật
            'auth_method': random.choice(['SSO', 'LDAP', 'OAuth', 'Username/Password']),
            'has_encryption': random.choice([True, False]),
            'has_audit_log': random.choice([True, False]),
            'security_level': random.randint(1, 5),
            'has_sec_docs': random.choice([True, False]),
            'compliance': 'ISO 27001, TCVN 11930',
            'notes_tab6': 'Ghi chú tab bảo mật',

            # Tab 7: Hạ tầng
            'server_config': random.choice(['Cloud VM', 'Physical Server', 'Container']),
            'storage_capacity': random.choice(['100GB - 1TB', '1TB - 10TB', '10TB - 100TB']),
            'backup_plan': random.choice(['Daily', 'Weekly', 'Real-time']),
            'dr_plan': random.choice(['Hot Standby', 'Warm Standby', 'Cold Backup']),
            'deploy_loc': DEPLOY_LOC[i % len(DEPLOY_LOC)],
            'compute_specs': '8 vCPU, 32GB RAM, 500GB SSD',
            'compute_type': COMPUTE_TYPE[i % len(COMPUTE_TYPE)],
            'deploy_freq': DEPLOY_FREQ[i % len(DEPLOY_FREQ)],
            'notes_tab7': 'Ghi chú tab hạ tầng',

            # Tab 8: Vận hành
            'responsible': f'Nguyễn Văn {chr(65 + i)}',
            'tech_admin': f'Trần Văn {chr(75 + i)}',
            'phone': f'0912{random.randint(100000, 999999)}',
            'email': f'contact{i}@example.gov.vn',
            'support_level': random.choice(['24/7', 'Business hours', 'On-demand']),
            'dev_type': DEV_TYPES[i % len(DEV_TYPES)],
            'developer': random.choice(['FPT', 'VNPT', 'Viettel', 'CMC']),
            'dev_team_size': random.randint(3, 15),
            'warranty': WARRANTY[i % len(WARRANTY)],
            'warranty_end': date(2025, random.randint(1, 12), random.randint(1, 28)),
            'has_maint': random.choice([True, False]),
            'maint_end': date(2026, random.randint(1, 12), random.randint(1, 28)),
            'operator': random.choice(['Nội bộ', 'FPT', 'VNPT']),
            'ops_team_size': random.randint(1, 5),
            'vendor_dep': VENDOR_DEP[i % len(VENDOR_DEP)],
            'can_self_maint': random.choice([True, False]),
            'notes_tab8': 'Ghi chú tab vận hành',

            # Tab 9: Đánh giá
            'int_readiness': ['Dễ chuẩn hóa', 'Có API tốt', 'Dữ liệu rõ nguồn gốc'],
            'blockers': ['Công nghệ cũ', 'Không có tài liệu'],
            'recommendation': RECOMMEND[i % len(RECOMMEND)],

            # Extra: Dates & versions
            'go_live': date(2020 + i % 5, (i * 3) % 12 + 1, (i * 7) % 28 + 1),
            'version': f'{random.randint(1, 5)}.{random.randint(0, 9)}.{random.randint(0, 9)}',
            'completion': random.randint(40, 100),

            # Level 2: Chi phí
            'initial_inv': random.randint(500, 5000) * 1000000,
            'dev_cost': random.randint(200, 2000) * 1000000,
            'license_annual': random.randint(50, 500) * 1000000,
            'maint_annual': random.randint(100, 500) * 1000000,
            'infra_annual': random.randint(50, 300) * 1000000,
            'personnel_annual': random.randint(200, 1000) * 1000000,
            'tco': random.randint(1000, 10000) * 1000000,
            'funding': random.choice(['Ngân sách nhà nước', 'Vốn ODA', 'Xã hội hóa']),

            # Level 2: Nhà cung cấp
            'vendor_name': random.choice(['FPT IS', 'VNPT-IT', 'Viettel Solutions', 'CMC TS']),
            'vendor_type': random.choice(['Tích hợp hệ thống', 'Phát triển PM', 'Cung cấp giải pháp']),
            'vendor_contact': f'Phạm Văn {chr(65 + i)}',
            'vendor_phone': f'0903{random.randint(100000, 999999)}',
            'vendor_email': f'vendor{i}@company.com',
            'contract_no': f'HĐ-2024-{str(i).zfill(3)}',
            'contract_start': date(2023, random.randint(1, 12), random.randint(1, 28)),
            'contract_end': date(2026, random.randint(1, 12), random.randint(1, 28)),
            'vendor_rating': random.randint(3, 5),
            'vendor_lockin': random.choice(['Cao', 'Trung bình', 'Thấp']),

            # Level 2: Hạ tầng chi tiết
            'num_servers': random.randint(1, 10),
            'server_specs': '8 vCPU, 32GB RAM, 500GB SSD',
            'total_cpu': random.randint(8, 64),
            'total_ram': random.randint(32, 256),
            'total_storage': random.randint(1, 50),
            'bandwidth': random.randint(100, 1000),
            'has_cdn': random.choice([True, False]),
            'has_lb': random.choice([True, False]),
            'backup_freq': random.choice(['Daily', 'Weekly', 'Real-time']),
            'backup_days': random.randint(7, 90),
            'has_dr': random.choice([True, False]),
            'rto': random.randint(1, 24),
            'rpo': random.randint(1, 12),

            # Level 2: Bảo mật chi tiết
            'auth_method_l2': random.choice(['SSO', 'LDAP', 'OAuth']),
            'has_mfa': random.choice([True, False]),
            'has_rbac': random.choice([True, False]),
            'enc_rest': random.choice([True, False]),
            'enc_transit': random.choice([True, False]),
            'has_fw': random.choice([True, False]),
            'has_waf': random.choice([True, False]),
            'has_ids': random.choice([True, False]),
            'has_av': random.choice([True, False]),
            'last_audit': date(2024, random.randint(1, 12), random.randint(1, 28)),
            'last_pentest': date(2024, random.randint(1, 12), random.randint(1, 28)),
            'has_vuln_scan': random.choice([True, False]),
            'sec_incidents': random.randint(0, 5),
            'sec_notes': 'Đã hoàn thành đánh giá bảo mật',
        }
        systems.append(sys)
    return systems

def create_sheet_tab1(wb, systems):
    """Sheet 1: Tab Cơ bản"""
    ws = wb.create_sheet("1. Cơ bản")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Tên tiếng Anh', 'Đơn vị',
        'Mô tả/Mục đích', 'Trạng thái', 'Mức độ quan trọng', 'Phạm vi',
        'Nhu cầu', 'Thời gian hoàn thành', 'Nhóm hệ thống',
        'Ngày vận hành', 'Phiên bản', '% Hoàn thành', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['name_en'], s['org'],
            s['purpose'], s['status'], s['criticality'], s['scope'],
            s['req_type'], date_str(s['target_date']), s['system_group'],
            date_str(s['go_live']), s['version'], f"{s['completion']}%", s['notes_tab1']
        ])

    # Apply status/criticality colors
    for row_idx in range(2, len(systems) + 2):
        status_cell = ws.cell(row=row_idx, column=7)
        if status_cell.value in STATUS_COLORS:
            status_cell.fill = STATUS_COLORS[status_cell.value]
        crit_cell = ws.cell(row=row_idx, column=8)
        if crit_cell.value in CRITICALITY_FONTS:
            crit_cell.font = CRITICALITY_FONTS[crit_cell.value]

    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab2(wb, systems):
    """Sheet 2: Tab Nghiệp vụ"""
    ws = wb.create_sheet("2. Nghiệp vụ")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Mục tiêu nghiệp vụ',
        'Quy trình nghiệp vụ', 'Có hồ sơ thiết kế', 'Đối tượng sử dụng',
        'Người dùng/năm', 'Tổng tài khoản', 'MAU', 'DAU', 'Số đơn vị', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], list_to_str(s['biz_objectives']),
            list_to_str(s['biz_processes']), bool_vn(s['has_design_docs']),
            list_to_str(s['user_types']), num_fmt(s['annual_users']),
            num_fmt(s['total_accounts']), num_fmt(s['mau']), num_fmt(s['dau']),
            s['num_orgs'], s['notes_tab2']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab3(wb, systems):
    """Sheet 3: Tab Kiến trúc & Công nghệ"""
    ws = wb.create_sheet("3. Kiến trúc")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Ngôn ngữ', 'Framework',
        'Database', 'Hosting', 'Backend Tech', 'Frontend Tech',
        'Kiểu kiến trúc', 'Container', 'Multi-tenant', 'Phân lớp',
        'API Style', 'Message Queue', 'Cache', 'Search Engine',
        'BI/Reporting', 'Source Repo', 'CI/CD', 'CI/CD Tool',
        'Auto Testing', 'Testing Tools', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['lang'], s['framework'],
            s['database'], s['hosting'], s['backend_tech'], s['frontend_tech'],
            s['arch_type'], s['containerization'], bool_vn(s['is_multi_tenant']),
            bool_vn(s['has_layered']), s['api_style'], s['msg_queue'],
            s['cache'], s['search'], s['bi_tool'], s['source_repo'],
            bool_vn(s['has_cicd']), s['cicd_tool'], bool_vn(s['has_auto_test']),
            s['test_tools'], s['notes_tab3']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab4(wb, systems):
    """Sheet 4: Tab Dữ liệu"""
    ws = wb.create_sheet("4. Dữ liệu")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Nguồn dữ liệu', 'Loại dữ liệu',
        'Phân loại DL', 'Khối lượng', 'Storage (GB)', 'File Storage (GB)',
        'Tăng trưởng (%)', 'Loại lưu trữ file', 'Số bản ghi', 'CSDL phụ',
        'Data Catalog', 'Ghi chú Catalog', 'MDM', 'Ghi chú MDM', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], list_to_str(s['data_sources']),
            list_to_str(s['data_types']), s['data_class'], s['data_volume'],
            s['storage_gb'], s['file_storage_gb'], f"{s['growth_rate']}%",
            s['file_storage_type'], num_fmt(s['record_count']), s['secondary_dbs'],
            bool_vn(s['has_catalog']), s['catalog_notes'],
            bool_vn(s['has_mdm']), s['mdm_notes'], s['notes_tab4']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab5(wb, systems):
    """Sheet 5: Tab Tích hợp"""
    ws = wb.create_sheet("5. Tích hợp")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Số API cung cấp', 'Số API sử dụng',
        'Chuẩn API', 'Có API Gateway', 'Tên Gateway', 'API Versioning',
        'Rate Limiting', 'Tài liệu API', 'Chuẩn version', 'Giám sát tích hợp',
        'HT nội bộ tích hợp', 'HT bên ngoài', 'Danh sách API',
        'Phương thức trao đổi', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['api_provided'], s['api_consumed'],
            s['api_standard'], bool_vn(s['has_api_gateway']), s['gateway_name'],
            bool_vn(s['has_versioning']), bool_vn(s['has_rate_limit']),
            s['api_docs'], s['version_std'], bool_vn(s['has_int_monitor']),
            list_to_str(s['internal_systems']), list_to_str(s['external_systems']),
            list_to_str(s['api_list']), s['exchange_method'], s['notes_tab5']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab6(wb, systems):
    """Sheet 6: Tab Bảo mật"""
    ws = wb.create_sheet("6. Bảo mật")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Phương thức xác thực',
        'Mã hóa dữ liệu', 'Audit Log', 'Cấp độ an toàn',
        'Có tài liệu ATTT', 'Tiêu chuẩn tuân thủ', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['auth_method'],
            bool_vn(s['has_encryption']), bool_vn(s['has_audit_log']),
            f"Cấp {s['security_level']}", bool_vn(s['has_sec_docs']),
            s['compliance'], s['notes_tab6']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab7(wb, systems):
    """Sheet 7: Tab Hạ tầng"""
    ws = wb.create_sheet("7. Hạ tầng")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Cấu hình máy chủ',
        'Dung lượng lưu trữ', 'Phương án sao lưu', 'Khôi phục thảm họa',
        'Vị trí triển khai', 'Cấu hình tính toán', 'Loại hạ tầng',
        'Tần suất triển khai', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['server_config'],
            s['storage_capacity'], s['backup_plan'], s['dr_plan'],
            s['deploy_loc'], s['compute_specs'], s['compute_type'],
            s['deploy_freq'], s['notes_tab7']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab8(wb, systems):
    """Sheet 8: Tab Vận hành"""
    ws = wb.create_sheet("8. Vận hành")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Người phụ trách', 'Quản trị KT',
        'SĐT', 'Email', 'Mức hỗ trợ', 'Loại phát triển', 'Đơn vị PT',
        'Team dev', 'Bảo hành', 'Hết bảo hành', 'Có HĐ bảo trì', 'Hết bảo trì',
        'Đơn vị vận hành', 'Team ops', 'Phụ thuộc NCC', 'Tự bảo trì', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['responsible'], s['tech_admin'],
            s['phone'], s['email'], s['support_level'], s['dev_type'],
            s['developer'], s['dev_team_size'], s['warranty'],
            date_str(s['warranty_end']), bool_vn(s['has_maint']),
            date_str(s['maint_end']), s['operator'], s['ops_team_size'],
            s['vendor_dep'], bool_vn(s['can_self_maint']), s['notes_tab8']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_tab9(wb, systems):
    """Sheet 9: Tab Đánh giá"""
    ws = wb.create_sheet("9. Đánh giá")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Điểm phù hợp tích hợp',
        'Điểm vướng mắc', 'Đề xuất'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], list_to_str(s['int_readiness']),
            list_to_str(s['blockers']), s['recommendation']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_cost(wb, systems):
    """Sheet 10: Chi phí (Level 2)"""
    ws = wb.create_sheet("10. Chi phí (L2)")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Đầu tư ban đầu (VNĐ)',
        'Chi phí PT (VNĐ)', 'License/năm (VNĐ)', 'Bảo trì/năm (VNĐ)',
        'Hạ tầng/năm (VNĐ)', 'Nhân sự/năm (VNĐ)', 'TCO (VNĐ)', 'Nguồn tài trợ'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'],
            f"{s['initial_inv']:,.0f}", f"{s['dev_cost']:,.0f}",
            f"{s['license_annual']:,.0f}", f"{s['maint_annual']:,.0f}",
            f"{s['infra_annual']:,.0f}", f"{s['personnel_annual']:,.0f}",
            f"{s['tco']:,.0f}", s['funding']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_vendor(wb, systems):
    """Sheet 11: Nhà cung cấp (Level 2)"""
    ws = wb.create_sheet("11. Nhà cung cấp (L2)")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Tên NCC', 'Loại NCC',
        'Người liên hệ', 'SĐT', 'Email', 'Số HĐ',
        'Bắt đầu HĐ', 'Kết thúc HĐ', 'Đánh giá', 'Rủi ro lock-in'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['vendor_name'], s['vendor_type'],
            s['vendor_contact'], s['vendor_phone'], s['vendor_email'],
            s['contract_no'], date_str(s['contract_start']),
            date_str(s['contract_end']), f"{s['vendor_rating']}/5", s['vendor_lockin']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_infra_l2(wb, systems):
    """Sheet 12: Hạ tầng chi tiết (Level 2)"""
    ws = wb.create_sheet("12. Hạ tầng CT (L2)")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Số máy chủ', 'Cấu hình server',
        'Tổng CPU', 'Tổng RAM (GB)', 'Tổng Storage (TB)', 'Bandwidth (Mbps)',
        'CDN', 'Load Balancer', 'Tần suất backup', 'Lưu trữ (ngày)',
        'DR', 'RTO (giờ)', 'RPO (giờ)'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['num_servers'], s['server_specs'],
            s['total_cpu'], s['total_ram'], s['total_storage'], s['bandwidth'],
            bool_vn(s['has_cdn']), bool_vn(s['has_lb']), s['backup_freq'],
            s['backup_days'], bool_vn(s['has_dr']), s['rto'], s['rpo']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def create_sheet_security_l2(wb, systems):
    """Sheet 13: Bảo mật chi tiết (Level 2)"""
    ws = wb.create_sheet("13. Bảo mật CT (L2)")
    headers = [
        'STT', 'Mã HT', 'Tên hệ thống', 'Xác thực', 'MFA', 'RBAC',
        'Mã hóa lưu trữ', 'Mã hóa truyền', 'Firewall', 'WAF', 'IDS/IPS',
        'Antivirus', 'Audit cuối', 'Pentest cuối', 'Quét lỗ hổng',
        'Sự cố bảo mật (năm)', 'Ghi chú'
    ]
    ws.append(headers)
    apply_header_style(ws)

    for s in systems:
        ws.append([
            s['stt'], s['code'], s['name'], s['auth_method_l2'],
            bool_vn(s['has_mfa']), bool_vn(s['has_rbac']),
            bool_vn(s['enc_rest']), bool_vn(s['enc_transit']),
            bool_vn(s['has_fw']), bool_vn(s['has_waf']), bool_vn(s['has_ids']),
            bool_vn(s['has_av']), date_str(s['last_audit']),
            date_str(s['last_pentest']), bool_vn(s['has_vuln_scan']),
            s['sec_incidents'], s['sec_notes']
        ])
    apply_data_style(ws)
    auto_fit_columns(ws)

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    systems = generate_sample_data(5)

    # Create all sheets (matching form tabs)
    create_sheet_tab1(wb, systems)    # Cơ bản
    create_sheet_tab2(wb, systems)    # Nghiệp vụ
    create_sheet_tab3(wb, systems)    # Kiến trúc
    create_sheet_tab4(wb, systems)    # Dữ liệu
    create_sheet_tab5(wb, systems)    # Tích hợp
    create_sheet_tab6(wb, systems)    # Bảo mật
    create_sheet_tab7(wb, systems)    # Hạ tầng
    create_sheet_tab8(wb, systems)    # Vận hành
    create_sheet_tab9(wb, systems)    # Đánh giá
    create_sheet_cost(wb, systems)    # Chi phí (L2)
    create_sheet_vendor(wb, systems)  # Nhà cung cấp (L2)
    create_sheet_infra_l2(wb, systems)   # Hạ tầng chi tiết (L2)
    create_sheet_security_l2(wb, systems) # Bảo mật chi tiết (L2)

    filename = f"SAMPLE-Danh-sach-He-thong-FULL-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = f"/Users/shimazu/Dropbox/9. active/consultant/support_b4t/thong_ke_he_thong/{filename}"
    wb.save(filepath)
    print(f"✓ Sample Excel FULL file created: {filepath}")
    print(f"  Total sheets: 13 (9 main tabs + 4 Level 2 tabs)")

if __name__ == "__main__":
    main()
