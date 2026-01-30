#!/usr/bin/env python3
"""
Test Excel Export from Systems Page with Playwright
- Login to the system
- Navigate to Systems page
- Export Excel file (click button, confirm modal)
- Check for formula injection bugs (values starting with =, +, -, @)
"""

import asyncio
import os
import tempfile
from datetime import datetime
from pathlib import Path

try:
    from playwright.async_api import async_playwright
    import openpyxl
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.check_call(["pip3", "install", "playwright", "openpyxl"])
    subprocess.check_call(["playwright", "install", "chromium"])
    from playwright.async_api import async_playwright
    import openpyxl

# Production URL
BASE_URL = "https://hientrangcds.mst.gov.vn"
# Test credentials (admin)
USERNAME = "admin"
PASSWORD = "Admin@2026"

# Download directory
DOWNLOAD_DIR = Path(tempfile.gettempdir()) / "excel_test"
DOWNLOAD_DIR.mkdir(exist_ok=True)


def check_excel_for_bugs(file_path: Path) -> list:
    """
    Check Excel file for formula injection bugs
    - Values starting with =, +, -, @ that aren't properly escaped
    """
    bugs = []

    try:
        wb = openpyxl.load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    value = cell.value
                    if value is not None and isinstance(value, str):
                        # Check for unescaped formula characters at start
                        if len(value) > 0:
                            first_char = value[0]
                            # If starts with =, +, -, @ and NOT escaped with '
                            if first_char in ['=', '+', '-', '@'] and not value.startswith("'"):
                                # Skip if it looks like a negative number
                                if first_char == '-' and len(value) > 1 and value[1:].replace('.', '').replace(',', '').isdigit():
                                    continue
                                # Skip common OK patterns
                                if value in ['-', '--', '---', '----']:
                                    continue

                                bugs.append({
                                    'sheet': sheet_name,
                                    'row': row_idx,
                                    'col': col_idx,
                                    'value': value[:100],
                                    'issue': f"Unescaped formula character: {first_char}"
                                })

        wb.close()
    except Exception as e:
        bugs.append({'error': str(e)})

    return bugs


async def test_excel_export():
    """Main test function"""
    print(f"\n{'='*60}")
    print("Excel Export Bug Test - Systems Page")
    print(f"{'='*60}\n")
    print(f"URL: {BASE_URL}")
    print(f"Download Dir: {DOWNLOAD_DIR}\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            accept_downloads=True,
            viewport={'width': 1920, 'height': 1080}
        )
        page = await context.new_page()

        try:
            # Step 1: Login
            print("Step 1: Logging in...")
            await page.goto(f"{BASE_URL}/login")
            await page.wait_for_load_state('networkidle')
            await asyncio.sleep(2)

            # Save screenshot of login page
            ss_path = DOWNLOAD_DIR / "login_page.png"
            await page.screenshot(path=ss_path)
            print(f"  (Login page screenshot: {ss_path.name})")

            # Find Ant Design inputs - try different selectors
            username_input = page.locator('input').first
            password_input = page.locator('input[type="password"]')

            await username_input.fill(USERNAME)
            await password_input.fill(PASSWORD)

            # Click login button
            login_btn = page.locator('button:has-text("Đăng nhập")')
            await login_btn.click()

            # Wait for login to complete
            await page.wait_for_load_state('networkidle')
            await asyncio.sleep(3)

            # Take screenshot after login
            ss_path = DOWNLOAD_DIR / "after_login.png"
            await page.screenshot(path=ss_path)
            print(f"  (After login screenshot: {ss_path.name})")

            # Check URL
            current_url = page.url
            print(f"  Current URL: {current_url}")

            if "/login" in current_url:
                print("  ✗ Login failed - still on login page")
                raise Exception("Login failed")

            print(f"  ✓ Logged in successfully\n")

            # Step 2: Navigate to Systems page via sidebar click
            print("Step 2: Navigating to Systems page...")

            # Try clicking menu item instead of direct URL navigation
            # Look for "Hệ thống" or "Systems" in sidebar
            systems_menu = page.locator('text=Hệ thống').first
            if await systems_menu.count() > 0:
                print("  Found 'Hệ thống' menu item, clicking...")
                await systems_menu.click()
                await page.wait_for_load_state('networkidle')
                await asyncio.sleep(3)
            else:
                # Try direct navigation
                print("  Menu item not found, trying direct URL...")
                await page.goto(f"{BASE_URL}/systems")
                await page.wait_for_load_state('networkidle')
                await asyncio.sleep(3)

            # Check if redirected to login
            if "/login" in page.url:
                print("  Session lost, trying to re-login...")
                # Fill login form again
                username_input = page.locator('input').first
                password_input = page.locator('input[type="password"]')
                await username_input.fill(USERNAME)
                await password_input.fill(PASSWORD)
                await page.locator('button:has-text("Đăng nhập")').click()
                await page.wait_for_load_state('networkidle')
                await asyncio.sleep(2)

                # Navigate again
                await page.goto(f"{BASE_URL}/systems")
                await page.wait_for_load_state('networkidle')
                await asyncio.sleep(3)

            # Take screenshot
            ss_path = DOWNLOAD_DIR / "systems_page.png"
            await page.screenshot(path=ss_path)
            print(f"  (Screenshot: {ss_path.name})")
            print(f"  Current URL: {page.url}")

            # Step 3: Click Export button
            print("\nStep 3: Finding and clicking export button...")

            # Debug: List all buttons
            buttons = page.locator('button')
            count = await buttons.count()
            print(f"  Found {count} buttons on page:")
            for i in range(min(count, 15)):
                btn = buttons.nth(i)
                btn_text = await btn.text_content()
                is_visible = await btn.is_visible()
                print(f"    Button {i+1}: '{btn_text[:50] if btn_text else '(empty)'}' visible={is_visible}")

            # Look for export button
            export_selectors = [
                'button:has-text("Xuất chi tiết")',
                'button:has-text("Xuất")',
                'button:has-text("Export")',
            ]

            export_btn = None
            for selector in export_selectors:
                btn = page.locator(selector).first
                if await btn.count() > 0 and await btn.is_visible():
                    export_btn = btn
                    print(f"\n  Found button with selector: {selector}")
                    break

            if not export_btn:
                raise Exception("Export button not found")

            await export_btn.click()
            await asyncio.sleep(1)

            # Step 4: Handle modal - confirm export
            print("\nStep 4: Handling export modal...")

            # Take screenshot of modal
            ss_path = DOWNLOAD_DIR / "export_modal.png"
            await page.screenshot(path=ss_path)
            print(f"  (Screenshot: {ss_path.name})")

            # Find confirm button in modal
            confirm_selectors = [
                '.ant-modal button:has-text("Xuất file")',
                '.ant-modal button:has-text("Xác nhận")',
                '.ant-modal button:has-text("OK")',
                '.ant-modal-footer button.ant-btn-primary',
                'button:has-text("Xuất file")',
            ]

            confirm_btn = None
            for selector in confirm_selectors:
                btn = page.locator(selector).first
                if await btn.count() > 0 and await btn.is_visible():
                    confirm_btn = btn
                    print(f"  Found confirm button with selector: {selector}")
                    break

            if not confirm_btn:
                print("  ! Confirm button not found in modal")
                raise Exception("Confirm button not found")

            # Wait for download
            async with page.expect_download(timeout=120000) as download_info:
                await confirm_btn.click()
                print("  Waiting for download...")

            download = await download_info.value
            download_file = DOWNLOAD_DIR / f"systems_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            await download.save_as(download_file)
            print(f"  ✓ Downloaded: {download_file.name}")

            # Step 5: Check for bugs
            print("\nStep 5: Checking Excel file for formula injection bugs...")
            bugs = check_excel_for_bugs(download_file)

            if bugs:
                print(f"  ✗ Found {len(bugs)} potential issues:")
                for bug in bugs[:30]:  # Show first 30
                    if 'error' in bug:
                        print(f"    - Error: {bug['error']}")
                    else:
                        print(f"    - Sheet '{bug['sheet']}' row {bug['row']} col {bug['col']}: {bug['value'][:50]}...")
            else:
                print("  ✓ No formula injection bugs found!")

            # Show sheet summary
            wb = openpyxl.load_workbook(download_file)
            print(f"\n  Excel file contains {len(wb.sheetnames)} sheets:")
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = sheet.max_row
                cols = sheet.max_column
                print(f"    - {sheet_name}: {rows} rows x {cols} cols")
            wb.close()

            print(f"\n{'='*60}")
            print("Test Complete!")
            print(f"{'='*60}")
            print(f"\nDownloaded files are in: {DOWNLOAD_DIR}")

        except Exception as e:
            print(f"\n✗ Test failed with error: {e}")
            # Take screenshot on error
            screenshot_path = DOWNLOAD_DIR / f"error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            await page.screenshot(path=screenshot_path)
            print(f"  Screenshot saved: {screenshot_path}")

        finally:
            await browser.close()


if __name__ == "__main__":
    asyncio.run(test_excel_export())
