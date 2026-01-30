#!/usr/bin/env python3
"""
Verify Excel Export Fix - Live Test
Tests that Sheet 3 (Danh sách HT) now has all 77 systems instead of only 20
"""

import requests
from openpyxl import load_workbook
import io
import sys

def test_excel_export():
    print("🧪 Live Testing Excel Export Fix...")
    print("=" * 60)

    # Login credentials
    login_url = "https://hientrangcds.mst.gov.vn/api/auth/login/"
    excel_url = "https://hientrangcds.mst.gov.vn/api/systems/export_excel/"

    # Use admin credentials
    credentials = {
        "username": "admin",
        "password": "admin123"  # Default admin password
    }

    session = requests.Session()

    try:
        # Step 1: Login
        print("📝 Step 1: Logging in...")
        response = session.post(login_url, json=credentials)

        if response.status_code == 200:
            print("✅ Login successful")
            token_data = response.json()
            token = token_data.get('access', '')
            session.headers.update({'Authorization': f'Bearer {token}'})
        else:
            print(f"❌ Login failed: {response.status_code}")
            print(f"Response: {response.text}")
            return False

        # Step 2: Download Excel
        print("\n📥 Step 2: Downloading Excel report...")
        response = session.get(excel_url)

        if response.status_code != 200:
            print(f"❌ Failed to download Excel: {response.status_code}")
            return False

        print("✅ Excel downloaded successfully")

        # Step 3: Analyze Excel
        print("\n🔍 Step 3: Analyzing Excel sheets...")
        excel_file = io.BytesIO(response.content)
        wb = load_workbook(excel_file)

        results = {}

        # Check all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_count = ws.max_row
            results[sheet_name] = row_count
            print(f"  📊 {sheet_name}: {row_count} rows")

        # Verify critical sheets
        print("\n" + "=" * 60)
        print("VERIFICATION RESULTS:")
        print("=" * 60)

        all_passed = True

        # Sheet 2: Theo đơn vị (32 organizations + 1 header = 33 rows)
        sheet2_rows = results.get('Theo đơn vị', 0)
        if sheet2_rows >= 33:
            print(f"✅ Sheet 2 (Theo đơn vị): {sheet2_rows} rows - PASS")
            print(f"   Expected: ≥33 (32 organizations + header)")
        else:
            print(f"❌ Sheet 2 (Theo đơn vị): {sheet2_rows} rows - FAIL")
            print(f"   Expected: ≥33, Got: {sheet2_rows}")
            all_passed = False

        # Sheet 3: Danh sách HT (77 systems + 1 header = 78 rows)
        sheet3_rows = results.get('Danh sách HT', 0)
        if sheet3_rows >= 78:
            print(f"✅ Sheet 3 (Danh sách HT): {sheet3_rows} rows - PASS")
            print(f"   Expected: ≥78 (77 systems + header)")
            systems_count = sheet3_rows - 1
            print(f"   ✓ All {systems_count} systems included!")
        elif sheet3_rows == 21:
            print(f"❌ Sheet 3 (Danh sách HT): {sheet3_rows} rows - FAIL")
            print(f"   Expected: ≥78, Got: {sheet3_rows}")
            print(f"   ⚠️  Bug still present: Only 20 systems (should be 77+)")
            all_passed = False
        else:
            print(f"⚠️  Sheet 3 (Danh sách HT): {sheet3_rows} rows - UNEXPECTED")
            print(f"   Expected: ≥78, Got: {sheet3_rows}")
            all_passed = False

        print("=" * 60)

        if all_passed:
            print("🎉 ALL TESTS PASSED - Fix is working!")
            return True
        else:
            print("❌ SOME TESTS FAILED - Fix may not be working")
            return False

    except Exception as e:
        print(f"❌ Error during test: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_excel_export()
    sys.exit(0 if success else 1)
